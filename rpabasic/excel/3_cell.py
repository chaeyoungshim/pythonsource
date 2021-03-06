from openpyxl import Workbook

wb = Workbook()
# 기본 시트 활성화
ws = wb.active
ws.title = "test"

# 셀에 데이터 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

# 셀 값 가져오기
print("A1", ws["A1"])
print("A1", ws["A1"].value)

# ws.cell(row, column)
print(ws.cell(1, 1).value)
print(ws.cell(1, 2).value)

# ws.cell(row,column)
c = ws.cell(row=1, column=3, value=10)  # 1행 3열의 value 를 뽑아와라
print(c.value)

wb.save("./rpabasic/excel/sample.xlsx")
